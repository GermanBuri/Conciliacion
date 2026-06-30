import csv
import logging
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.models.bank_transaction import BankTransaction
from app.models.ledger_transaction import LedgerTransaction
from app.models.uploaded_file import UploadedFile
from app.models.user import User
from app.repositories.bank_transaction import BankTransactionRepository
from app.repositories.import_batch import ImportBatchRepository
from app.repositories.ledger_transaction import LedgerTransactionRepository
from app.repositories.uploaded_file import UploadedFileRepository
from app.schemas.imports import ImportBatchListResponse

logger = logging.getLogger(__name__)

IMPORT_STATUS_UPLOADED = "uploaded"
IMPORT_STATUS_PROCESSING = "processing"
IMPORT_STATUS_VALIDATED = "validated"
IMPORT_STATUS_NORMALIZED = "normalized"
IMPORT_STATUS_FAILED = "failed"

BATCH_TYPE_BANK = "bank"
BATCH_TYPE_LEDGER = "ledger"

FILE_TYPE_TO_BATCH_TYPE = {
    "bank_statement": BATCH_TYPE_BANK,
    "ledger": BATCH_TYPE_LEDGER,
}

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}

BANK_COLUMN_ALIASES = {
    "transaction_date": {"transaction_date", "date", "fecha", "posting_date", "fecha_transaccion"},
    "description": {"description", "descripcion", "detail", "detalle", "concepto"},
    "reference": {"reference", "referencia", "ref"},
    "amount": {"amount", "importe", "monto", "valor"},
    "balance": {"balance", "saldo", "running_balance"},
    "document_number": {"document_number", "documento", "numero_documento", "document", "doc_number"},
}

LEDGER_COLUMN_ALIASES = {
    "transaction_date": {"transaction_date", "date", "fecha", "posting_date", "fecha_transaccion"},
    "account_code": {"account_code", "codigo_cuenta", "cuenta", "codigo", "account"},
    "account_name": {"account_name", "nombre_cuenta", "descripcion_cuenta", "account_name_desc"},
    "third_party": {"third_party", "tercero", "beneficiario", "counterparty"},
    "document_number": {"document_number", "documento", "numero_documento", "comprobante", "doc_number"},
    "debit": {"debit", "debito", "cargo"},
    "credit": {"credit", "credito", "abono"},
}

REQUIRED_BANK_COLUMNS = {"transaction_date", "description", "amount"}
REQUIRED_LEDGER_COLUMNS = {"transaction_date", "account_code", "account_name", "debit", "credit"}


@dataclass
class ParsedImportResult:
    total_records: int
    valid_records: int
    invalid_records: int
    bank_transactions: list[BankTransaction]
    ledger_transactions: list[LedgerTransaction]


class ImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.batch_repository = ImportBatchRepository(db)
        self.uploaded_file_repository = UploadedFileRepository(db)
        self.bank_transaction_repository = BankTransactionRepository(db)
        self.ledger_transaction_repository = LedgerTransactionRepository(db)

    def process_file(self, *, file_id: uuid.UUID, current_user: User):
        uploaded_file = self._get_uploaded_file(file_id=file_id, current_user=current_user)
        batch_type = self._resolve_batch_type(uploaded_file.file_type)
        self._validate_extension(uploaded_file.extension)
        self._ensure_file_exists(uploaded_file.storage_path)
        self._validate_reprocessing(file_id)

        import_batch = self.batch_repository.create(
            company_id=uploaded_file.company_id or current_user.company_id,
            file_id=uploaded_file.id,
            batch_type=batch_type,
            status=IMPORT_STATUS_PROCESSING,
        )
        self.uploaded_file_repository.update_processing_status(
            uploaded_file, processing_status=IMPORT_STATUS_PROCESSING
        )

        try:
            rows = self._read_rows(Path(uploaded_file.storage_path), uploaded_file.extension)
            parsed_result = self._parse_rows(rows=rows, batch_id=import_batch.id, batch_type=batch_type)
            import_batch = self.batch_repository.update_counts_and_status(
                import_batch,
                status=IMPORT_STATUS_VALIDATED,
                total_records=parsed_result.total_records,
                valid_records=parsed_result.valid_records,
                invalid_records=parsed_result.invalid_records,
            )

            if batch_type == BATCH_TYPE_BANK and parsed_result.bank_transactions:
                self.bank_transaction_repository.bulk_create(parsed_result.bank_transactions)
            if batch_type == BATCH_TYPE_LEDGER and parsed_result.ledger_transactions:
                self.ledger_transaction_repository.bulk_create(parsed_result.ledger_transactions)

            import_batch = self.batch_repository.update_status(
                import_batch,
                status=IMPORT_STATUS_NORMALIZED,
            )
            self.uploaded_file_repository.update_processing_status(
                uploaded_file,
                processing_status=IMPORT_STATUS_NORMALIZED,
            )
            return import_batch
        except HTTPException as exc:
            logger.exception("Import validation failed for file_id=%s", file_id)
            self.db.rollback()
            self.batch_repository.update_status(import_batch, status=IMPORT_STATUS_FAILED)
            self.uploaded_file_repository.update_processing_status(
                uploaded_file,
                processing_status=IMPORT_STATUS_FAILED,
            )
            raise exc
        except (OSError, SQLAlchemyError, ValueError) as exc:
            logger.exception("Import processing failed for file_id=%s", file_id)
            self.db.rollback()
            self.batch_repository.update_status(import_batch, status=IMPORT_STATUS_FAILED)
            self.uploaded_file_repository.update_processing_status(
                uploaded_file,
                processing_status=IMPORT_STATUS_FAILED,
            )
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="No fue posible procesar el archivo importado.",
            ) from exc

    def list_batches(
        self,
        *,
        company_id: uuid.UUID | None,
        skip: int = 0,
        limit: int = 50,
    ) -> ImportBatchListResponse:
        items, total = self.batch_repository.list_batches(
            company_id=company_id,
            skip=skip,
            limit=limit,
        )
        return ImportBatchListResponse(items=items, total=total)

    def get_batch(self, *, batch_id: uuid.UUID, company_id: uuid.UUID | None):
        import_batch = self.batch_repository.get_by_id(batch_id, company_id=company_id)
        if import_batch is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Lote de importacion no encontrado.",
            )
        return import_batch

    def _get_uploaded_file(self, *, file_id: uuid.UUID, current_user: User) -> UploadedFile:
        uploaded_file = self.uploaded_file_repository.get_by_id(
            file_id,
            company_id=current_user.company_id,
        )
        if uploaded_file is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Archivo cargado no encontrado.",
            )
        return uploaded_file

    @staticmethod
    def _resolve_batch_type(file_type: str) -> str:
        batch_type = FILE_TYPE_TO_BATCH_TYPE.get(file_type)
        if batch_type is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo indicado no puede procesarse como importacion.",
            )
        return batch_type

    @staticmethod
    def _validate_extension(extension: str) -> None:
        if extension.lower() not in SUPPORTED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Solo se soportan archivos CSV y XLSX para esta fase.",
            )

    @staticmethod
    def _ensure_file_exists(storage_path: str) -> None:
        if not Path(storage_path).exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="El archivo almacenado no existe en la ruta esperada.",
            )

    def _validate_reprocessing(self, file_id: uuid.UUID) -> None:
        latest_batch = self.batch_repository.get_latest_by_file_id(file_id)
        if latest_batch and latest_batch.status != IMPORT_STATUS_FAILED:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="El archivo ya fue procesado o tiene un lote activo.",
            )

    def _read_rows(self, file_path: Path, extension: str) -> list[dict[str, object]]:
        if extension.lower() == ".csv":
            return self._read_csv(file_path)
        if extension.lower() == ".xlsx":
            return self._read_xlsx(file_path)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Extension de archivo no soportada para importacion.",
        )

    def _read_csv(self, file_path: Path) -> list[dict[str, object]]:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with file_path.open("r", encoding=encoding, newline="") as csv_file:
                    reader = csv.DictReader(csv_file)
                    self._ensure_headers(reader.fieldnames)
                    return [self._sanitize_row(row) for row in reader if row]
            except UnicodeDecodeError:
                continue

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No fue posible decodificar el archivo CSV.",
        )

    def _read_xlsx(self, file_path: Path) -> list[dict[str, object]]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="El archivo no contiene filas para importar.",
                )

            headers = [self._sanitize_header(value) for value in rows[0]]
            self._ensure_headers(headers)
            items: list[dict[str, object]] = []
            for raw_row in rows[1:]:
                row = {
                    headers[index]: raw_row[index] if index < len(raw_row) else None
                    for index in range(len(headers))
                }
                items.append(self._sanitize_row(row))
            return items
        finally:
            workbook.close()

    @staticmethod
    def _ensure_headers(headers: list[str] | tuple[str, ...] | None) -> None:
        if not headers or not any(header for header in headers):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo no contiene encabezados validos.",
            )

    def _parse_rows(
        self,
        *,
        rows: list[dict[str, object]],
        batch_id: uuid.UUID,
        batch_type: str,
    ) -> ParsedImportResult:
        if batch_type == BATCH_TYPE_BANK:
            return self._parse_bank_rows(rows=rows, batch_id=batch_id)
        return self._parse_ledger_rows(rows=rows, batch_id=batch_id)

    def _parse_bank_rows(self, *, rows: list[dict[str, object]], batch_id: uuid.UUID) -> ParsedImportResult:
        mapped_rows = self._map_columns(rows, BANK_COLUMN_ALIASES, REQUIRED_BANK_COLUMNS)
        valid_items: list[BankTransaction] = []
        invalid_records = 0

        for index, row in enumerate(mapped_rows, start=2):
            try:
                transaction = BankTransaction(
                    batch_id=batch_id,
                    transaction_date=self._normalize_date(row["transaction_date"]),
                    description=self._normalize_text(row["description"], "description"),
                    reference=self._normalize_optional_text(row.get("reference")),
                    amount=self._normalize_money(row["amount"], "amount"),
                    balance=self._normalize_optional_money(row.get("balance")),
                    document_number=self._normalize_optional_text(row.get("document_number")),
                    raw_data=self._raw_data_for_storage(row),
                )
                valid_items.append(transaction)
            except ValueError as exc:
                invalid_records += 1
                logger.warning("Bank import row %s skipped: %s", index, exc)

        return ParsedImportResult(
            total_records=len(mapped_rows),
            valid_records=len(valid_items),
            invalid_records=invalid_records,
            bank_transactions=valid_items,
            ledger_transactions=[],
        )

    def _parse_ledger_rows(self, *, rows: list[dict[str, object]], batch_id: uuid.UUID) -> ParsedImportResult:
        mapped_rows = self._map_columns(rows, LEDGER_COLUMN_ALIASES, REQUIRED_LEDGER_COLUMNS)
        valid_items: list[LedgerTransaction] = []
        invalid_records = 0

        for index, row in enumerate(mapped_rows, start=2):
            try:
                transaction = LedgerTransaction(
                    batch_id=batch_id,
                    transaction_date=self._normalize_date(row["transaction_date"]),
                    account_code=self._normalize_text(row["account_code"], "account_code"),
                    account_name=self._normalize_text(row["account_name"], "account_name"),
                    third_party=self._normalize_optional_text(row.get("third_party")),
                    document_number=self._normalize_optional_text(row.get("document_number")),
                    debit=self._normalize_money(row["debit"], "debit"),
                    credit=self._normalize_money(row["credit"], "credit"),
                    raw_data=self._raw_data_for_storage(row),
                )
                valid_items.append(transaction)
            except ValueError as exc:
                invalid_records += 1
                logger.warning("Ledger import row %s skipped: %s", index, exc)

        return ParsedImportResult(
            total_records=len(mapped_rows),
            valid_records=len(valid_items),
            invalid_records=invalid_records,
            bank_transactions=[],
            ledger_transactions=valid_items,
        )

    def _map_columns(
        self,
        rows: list[dict[str, object]],
        aliases: dict[str, set[str]],
        required_columns: set[str],
    ) -> list[dict[str, object]]:
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo no contiene registros para importar.",
            )

        sample_headers = {self._sanitize_header(key) for key in rows[0].keys()}
        header_map = self._build_header_map(sample_headers, aliases)
        missing_columns = sorted(required_columns - set(header_map.keys()))
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Faltan columnas obligatorias: {', '.join(missing_columns)}.",
            )

        mapped_rows: list[dict[str, object]] = []
        for row in rows:
            normalized_row = {self._sanitize_header(key): value for key, value in row.items()}
            mapped_rows.append(
                {
                    canonical_name: normalized_row.get(source_name)
                    for canonical_name, source_name in header_map.items()
                }
            )
        return mapped_rows

    def _build_header_map(
        self,
        headers: set[str],
        aliases: dict[str, set[str]],
    ) -> dict[str, str]:
        header_map: dict[str, str] = {}
        for canonical_name, options in aliases.items():
            matched = next((header for header in headers if header in options), None)
            if matched is not None:
                header_map[canonical_name] = matched
        return header_map

    @staticmethod
    def _sanitize_row(row: dict[str, object]) -> dict[str, object]:
        return {
            ImportService._sanitize_header(key): value
            for key, value in row.items()
            if key is not None and str(key).strip()
        }

    @staticmethod
    def _sanitize_header(value: object) -> str:
        return str(value or "").strip().lower().replace(" ", "_")

    @staticmethod
    def _normalize_text(value: object, field_name: str) -> str:
        normalized = str(value or "").strip()
        if not normalized:
            raise ValueError(f"El campo {field_name} es obligatorio.")
        return normalized

    @staticmethod
    def _normalize_optional_text(value: object) -> str | None:
        normalized = str(value or "").strip()
        return normalized or None

    @staticmethod
    def _normalize_date(value: object) -> date:
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)):
            raise ValueError("No fue posible interpretar la fecha numerica.")

        normalized = str(value or "").strip()
        if not normalized:
            raise ValueError("La fecha es obligatoria.")

        date_formats = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d")
        for date_format in date_formats:
            try:
                return datetime.strptime(normalized, date_format).date()
            except ValueError:
                continue
        raise ValueError(f"Formato de fecha invalido: {normalized}.")

    @staticmethod
    def _normalize_money(value: object, field_name: str) -> Decimal:
        normalized = ImportService._normalize_decimal(value)
        if normalized is None:
            raise ValueError(f"El campo {field_name} es obligatorio.")
        return normalized

    @staticmethod
    def _normalize_optional_money(value: object) -> Decimal | None:
        return ImportService._normalize_decimal(value)

    @staticmethod
    def _normalize_decimal(value: object) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value.quantize(Decimal("0.01"))
        if isinstance(value, (int, float)):
            return Decimal(str(value)).quantize(Decimal("0.01"))

        normalized = str(value).strip()
        if not normalized:
            return None

        normalized = normalized.replace(" ", "")
        if "," in normalized and "." in normalized:
            if normalized.rfind(",") > normalized.rfind("."):
                normalized = normalized.replace(".", "").replace(",", ".")
            else:
                normalized = normalized.replace(",", "")
        elif "," in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")

        try:
            return Decimal(normalized).quantize(Decimal("0.01"))
        except InvalidOperation as exc:
            raise ValueError(f"Valor monetario invalido: {value}.") from exc

    @staticmethod
    def _raw_data_for_storage(row: dict[str, object]) -> dict[str, object]:
        serialized: dict[str, object] = {}
        for key, value in row.items():
            if isinstance(value, datetime):
                serialized[key] = value.isoformat()
            elif isinstance(value, date):
                serialized[key] = value.isoformat()
            elif isinstance(value, Decimal):
                serialized[key] = str(value)
            else:
                serialized[key] = value
        return serialized
