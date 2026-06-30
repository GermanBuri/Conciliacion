import re
import tempfile
import uuid
from dataclasses import dataclass
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.bank_transaction import BankTransaction
from app.models.import_batch import ImportBatch
from app.models.ledger_transaction import LedgerTransaction
from app.models.reconciliation_audit_log import ReconciliationAuditLog
from app.models.reconciliation_match import ReconciliationMatch
from app.models.reconciliation_run import ReconciliationRun
from app.models.user import User
from app.repositories.reconciliation import (
    ReconciliationAuditLogRepository,
    ReconciliationMatchRepository,
    ReconciliationRunRepository,
)
from app.schemas.reconciliation import (
    ReconciliationAuditLogListResponse,
    ReconciliationManualMatchRequest,
    ReconciliationRunDetail,
    ReconciliationRunListResponse,
    ReconciliationRunRead,
    ReconciliationRunRequest,
    ReconciliationUnmatchRequest,
)

BATCH_TYPE_BANK = "bank"
BATCH_TYPE_LEDGER = "ledger"
MATCHED = "matched"
UNMATCHED_BANK = "unmatched_bank"
UNMATCHED_LEDGER = "unmatched_ledger"
POSSIBLE_MATCH = "possible_match"
DESCRIPTION_MATCH_THRESHOLD = Decimal("0.60")
POSSIBLE_MATCH_MAX_AMOUNT_DIFFERENCE = Decimal("1000.00")
POSSIBLE_MATCH_MAX_DAYS_DIFFERENCE = 3


@dataclass
class CandidateMatch:
    bank_transaction: BankTransaction
    ledger_transaction: LedgerTransaction
    status: str
    amount_difference: Decimal
    days_difference: int
    description_similarity: Decimal
    match_score: int
    priority: int
    notes: str | None


@dataclass
class MatchDraft:
    bank_transaction_id: uuid.UUID | None
    ledger_transaction_id: uuid.UUID | None
    status: str
    match_score: int
    amount_difference: Decimal | None
    days_difference: int | None
    description_similarity: Decimal | None
    notes: str | None


class ReconciliationService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.audit_repository = ReconciliationAuditLogRepository(db)
        self.run_repository = ReconciliationRunRepository(db)
        self.match_repository = ReconciliationMatchRepository(db)

    def run(self, *, payload: ReconciliationRunRequest, current_user: User) -> ReconciliationRunDetail:
        bank_batch = self._get_batch(
            batch_id=payload.bank_batch_id,
            company_id=current_user.company_id,
            expected_batch_type=BATCH_TYPE_BANK,
        )
        ledger_batch = self._get_batch(
            batch_id=payload.ledger_batch_id,
            company_id=current_user.company_id,
            expected_batch_type=BATCH_TYPE_LEDGER,
        )
        if bank_batch.company_id != ledger_batch.company_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Los lotes no pertenecen a la misma empresa.",
            )

        bank_transactions = self._list_bank_transactions(bank_batch.id)
        ledger_transactions = self._list_ledger_transactions(ledger_batch.id)
        matches, summary = self._build_matches(
            bank_transactions=bank_transactions,
            ledger_transactions=ledger_transactions,
            tolerance_amount=payload.tolerance_amount,
            tolerance_days=payload.tolerance_days,
        )

        run = ReconciliationRun(
            company_id=bank_batch.company_id,
            bank_batch_id=bank_batch.id,
            ledger_batch_id=ledger_batch.id,
            created_by_id=current_user.id,
            tolerance_amount=payload.tolerance_amount,
            tolerance_days=payload.tolerance_days,
            status="completed",
            total_bank=summary["total_bank"],
            total_ledger=summary["total_ledger"],
            matched_count=summary["matched_count"],
            unmatched_bank_count=summary["unmatched_bank_count"],
            unmatched_ledger_count=summary["unmatched_ledger_count"],
            possible_match_count=summary["possible_match_count"],
        )
        run = self.run_repository.create(run)

        persisted_matches = [
            ReconciliationMatch(
                run_id=run.id,
                bank_transaction_id=item.bank_transaction_id,
                ledger_transaction_id=item.ledger_transaction_id,
                status=item.status,
                match_score=item.match_score,
                amount_difference=item.amount_difference,
                days_difference=item.days_difference,
                description_similarity=item.description_similarity,
                notes=item.notes,
            )
            for item in matches
        ]
        if persisted_matches:
            self.match_repository.bulk_create(persisted_matches)

        self._create_audit_log(
            run_id=run.id,
            user_id=current_user.id,
            action="run_created",
            description="Conciliación automática ejecutada",
            metadata_json={
                "bank_batch_id": str(run.bank_batch_id),
                "ledger_batch_id": str(run.ledger_batch_id),
                "tolerance_amount": str(run.tolerance_amount),
                "tolerance_days": run.tolerance_days,
            },
        )

        return self.get_run(run_id=run.id, company_id=current_user.company_id)

    def list_runs(
        self,
        *,
        company_id: uuid.UUID | None,
        skip: int = 0,
        limit: int = 50,
    ) -> ReconciliationRunListResponse:
        items, total = self.run_repository.list_runs(company_id=company_id, skip=skip, limit=limit)
        return ReconciliationRunListResponse(items=items, total=total)

    def get_run(self, *, run_id: uuid.UUID, company_id: uuid.UUID | None) -> ReconciliationRunDetail:
        run = self.run_repository.get_by_id(run_id, company_id=company_id)
        if run is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ejecucion de conciliacion no encontrada.",
            )
        matches = self.match_repository.list_by_run_id(run.id)
        run_data = ReconciliationRunRead.model_validate(run).model_dump()
        return ReconciliationRunDetail(**run_data, matches=matches)

    def list_audit_logs(self, *, run_id: uuid.UUID, company_id: uuid.UUID | None) -> ReconciliationAuditLogListResponse:
        run = self._get_run_or_404(run_id=run_id, company_id=company_id)
        items, total = self.audit_repository.list_by_run_id(run.id)
        return ReconciliationAuditLogListResponse(items=list(items), total=total)

    def export_run(
        self,
        *,
        run_id: uuid.UUID,
        company_id: uuid.UUID | None,
        user_id: uuid.UUID | None = None,
    ) -> tuple[Path, str]:
        run = self.run_repository.get_by_id(run_id, company_id=company_id)
        if run is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ejecucion de conciliacion no encontrada.",
            )

        matches = list(self.match_repository.list_by_run_id(run.id))
        bank_ids = [item.bank_transaction_id for item in matches if item.bank_transaction_id is not None]
        ledger_ids = [item.ledger_transaction_id for item in matches if item.ledger_transaction_id is not None]
        bank_transactions = self._get_bank_transactions_map(bank_ids)
        ledger_transactions = self._get_ledger_transactions_map(ledger_ids)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumen"
        matched_sheet = workbook.create_sheet("Matched")
        unmatched_bank_sheet = workbook.create_sheet("Unmatched Bank")
        unmatched_ledger_sheet = workbook.create_sheet("Unmatched Ledger")

        self._populate_summary_sheet(summary_sheet, run)
        self._populate_matched_sheet(matched_sheet, matches, bank_transactions, ledger_transactions)
        self._populate_unmatched_bank_sheet(unmatched_bank_sheet, matches, bank_transactions)
        self._populate_unmatched_ledger_sheet(unmatched_ledger_sheet, matches, ledger_transactions)

        export_dir = settings.upload_path.parent / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        filename = f"Conciliacion_{run.id}.xlsx"
        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx",
            prefix=f"reconciliation_{run.id}_",
            dir=export_dir,
        ) as temp_file:
            temp_path = Path(temp_file.name)

        workbook.save(temp_path)
        workbook.close()
        self._create_audit_log(
            run_id=run.id,
            user_id=user_id,
            action="export",
            description="Conciliación exportada a Excel",
            metadata_json={
                "filename": filename,
                "path": str(temp_path),
            },
        )
        return temp_path, filename

    def manual_match(
        self,
        *,
        run_id: uuid.UUID,
        payload: ReconciliationManualMatchRequest,
        company_id: uuid.UUID | None,
        user_id: uuid.UUID | None,
    ) -> ReconciliationRunDetail:
        run = self._get_run_or_404(run_id=run_id, company_id=company_id)
        bank_transaction = self._get_bank_transaction_for_run(
            transaction_id=payload.bank_transaction_id,
            expected_batch_id=run.bank_batch_id,
        )
        ledger_transaction = self._get_ledger_transaction_for_run(
            transaction_id=payload.ledger_transaction_id,
            expected_batch_id=run.ledger_batch_id,
        )
        matches = list(self.match_repository.list_by_run_id(run.id))
        self._ensure_transaction_not_already_matched(
            matches=matches,
            bank_transaction_id=bank_transaction.id,
            ledger_transaction_id=ledger_transaction.id,
        )

        matches_to_replace = [
            match
            for match in matches
            if match.bank_transaction_id == bank_transaction.id or match.ledger_transaction_id == ledger_transaction.id
        ]
        if matches_to_replace:
            self.match_repository.delete_many(matches_to_replace)

        manual_match = ReconciliationMatch(
            run_id=run.id,
            bank_transaction_id=bank_transaction.id,
            ledger_transaction_id=ledger_transaction.id,
            status=MATCHED,
            match_score=100,
            amount_difference=abs(bank_transaction.amount - self._ledger_amount(ledger_transaction)),
            days_difference=abs((bank_transaction.transaction_date - ledger_transaction.transaction_date).days),
            description_similarity=self._description_similarity(bank_transaction, ledger_transaction),
            notes=(payload.notes or "Conciliación manual").strip() or "Conciliación manual",
        )
        self.match_repository.create(manual_match)
        self._recalculate_run_counts(run)
        self._create_audit_log(
            run_id=run.id,
            user_id=user_id,
            action="manual_match",
            description="Conciliación manual realizada",
            metadata_json={
                "bank_transaction_id": str(bank_transaction.id),
                "ledger_transaction_id": str(ledger_transaction.id),
                "notes": manual_match.notes,
            },
        )
        return self.get_run(run_id=run.id, company_id=company_id)

    def unmatch(
        self,
        *,
        run_id: uuid.UUID,
        payload: ReconciliationUnmatchRequest,
        company_id: uuid.UUID | None,
        user_id: uuid.UUID | None,
    ) -> ReconciliationRunDetail:
        run = self._get_run_or_404(run_id=run_id, company_id=company_id)
        match = self.match_repository.get_by_id(payload.match_id, run_id=run.id)
        if match is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Match de conciliacion no encontrado.",
            )
        if match.status != MATCHED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Solo se pueden revertir matches en estado matched.",
            )

        bank_transaction = None
        ledger_transaction = None
        if match.bank_transaction_id is not None:
            bank_transaction = self._get_bank_transaction_for_run(
                transaction_id=match.bank_transaction_id,
                expected_batch_id=run.bank_batch_id,
            )
        if match.ledger_transaction_id is not None:
            ledger_transaction = self._get_ledger_transaction_for_run(
                transaction_id=match.ledger_transaction_id,
                expected_batch_id=run.ledger_batch_id,
            )

        reversal_note = (payload.notes or "Se reversa conciliación manual").strip() or "Se reversa conciliación manual"
        self.match_repository.delete(match)

        recreated_matches: list[ReconciliationMatch] = []
        if bank_transaction is not None:
            recreated_matches.append(
                ReconciliationMatch(
                    run_id=run.id,
                    bank_transaction_id=bank_transaction.id,
                    ledger_transaction_id=None,
                    status=UNMATCHED_BANK,
                    match_score=0,
                    amount_difference=None,
                    days_difference=None,
                    description_similarity=None,
                    notes=reversal_note,
                )
            )
        if ledger_transaction is not None:
            recreated_matches.append(
                ReconciliationMatch(
                    run_id=run.id,
                    bank_transaction_id=None,
                    ledger_transaction_id=ledger_transaction.id,
                    status=UNMATCHED_LEDGER,
                    match_score=0,
                    amount_difference=None,
                    days_difference=None,
                    description_similarity=None,
                    notes=reversal_note,
                )
            )
        if recreated_matches:
            self.match_repository.bulk_create(recreated_matches)

        self._recalculate_run_counts(run)
        self._create_audit_log(
            run_id=run.id,
            user_id=user_id,
            action="unmatch",
            description="Conciliación reversada",
            metadata_json={
                "match_id": str(payload.match_id),
                "notes": reversal_note,
                "bank_transaction_id": str(bank_transaction.id) if bank_transaction is not None else None,
                "ledger_transaction_id": str(ledger_transaction.id) if ledger_transaction is not None else None,
            },
        )
        return self.get_run(run_id=run.id, company_id=company_id)

    def _create_audit_log(
        self,
        *,
        run_id: uuid.UUID,
        user_id: uuid.UUID | None,
        action: str,
        description: str,
        metadata_json: dict | None = None,
    ) -> ReconciliationAuditLog:
        audit_log = ReconciliationAuditLog(
            run_id=run_id,
            user_id=user_id,
            action=action,
            description=description,
            metadata_json=metadata_json,
        )
        return self.audit_repository.create(audit_log)

    def _get_batch(
        self,
        *,
        batch_id: uuid.UUID,
        company_id: uuid.UUID | None,
        expected_batch_type: str,
    ) -> ImportBatch:
        query = select(ImportBatch).where(ImportBatch.id == batch_id)
        if company_id is not None:
            query = query.where(ImportBatch.company_id == company_id)
        batch = self.db.execute(query).scalar_one_or_none()
        if batch is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Lote de importacion no encontrado.",
            )
        if batch.batch_type != expected_batch_type:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El lote indicado no corresponde al tipo esperado para la conciliacion.",
            )
        return batch

    def _list_bank_transactions(self, batch_id: uuid.UUID) -> list[BankTransaction]:
        query = (
            select(BankTransaction)
            .where(BankTransaction.batch_id == batch_id)
            .order_by(BankTransaction.transaction_date.asc(), BankTransaction.id.asc())
        )
        return list(self.db.execute(query).scalars().all())

    def _list_ledger_transactions(self, batch_id: uuid.UUID) -> list[LedgerTransaction]:
        query = (
            select(LedgerTransaction)
            .where(LedgerTransaction.batch_id == batch_id)
            .order_by(LedgerTransaction.transaction_date.asc(), LedgerTransaction.id.asc())
        )
        return list(self.db.execute(query).scalars().all())

    def _get_run_or_404(self, *, run_id: uuid.UUID, company_id: uuid.UUID | None) -> ReconciliationRun:
        run = self.run_repository.get_by_id(run_id, company_id=company_id)
        if run is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ejecucion de conciliacion no encontrada.",
            )
        return run

    def _get_bank_transaction_for_run(
        self,
        *,
        transaction_id: uuid.UUID,
        expected_batch_id: uuid.UUID,
    ) -> BankTransaction:
        query = select(BankTransaction).where(
            BankTransaction.id == transaction_id,
            BankTransaction.batch_id == expected_batch_id,
        )
        transaction = self.db.execute(query).scalar_one_or_none()
        if transaction is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Movimiento bancario no encontrado para esta conciliacion.",
            )
        return transaction

    def _get_ledger_transaction_for_run(
        self,
        *,
        transaction_id: uuid.UUID,
        expected_batch_id: uuid.UUID,
    ) -> LedgerTransaction:
        query = select(LedgerTransaction).where(
            LedgerTransaction.id == transaction_id,
            LedgerTransaction.batch_id == expected_batch_id,
        )
        transaction = self.db.execute(query).scalar_one_or_none()
        if transaction is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Movimiento contable no encontrado para esta conciliacion.",
            )
        return transaction

    def _get_bank_transactions_map(self, transaction_ids: list[uuid.UUID]) -> dict[uuid.UUID, BankTransaction]:
        if not transaction_ids:
            return {}
        query = select(BankTransaction).where(BankTransaction.id.in_(transaction_ids))
        items = self.db.execute(query).scalars().all()
        return {item.id: item for item in items}

    def _get_ledger_transactions_map(self, transaction_ids: list[uuid.UUID]) -> dict[uuid.UUID, LedgerTransaction]:
        if not transaction_ids:
            return {}
        query = select(LedgerTransaction).where(LedgerTransaction.id.in_(transaction_ids))
        items = self.db.execute(query).scalars().all()
        return {item.id: item for item in items}

    def _ensure_transaction_not_already_matched(
        self,
        *,
        matches: list[ReconciliationMatch],
        bank_transaction_id: uuid.UUID,
        ledger_transaction_id: uuid.UUID,
    ) -> None:
        for match in matches:
            if match.status != MATCHED:
                continue
            if match.bank_transaction_id == bank_transaction_id:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="El movimiento bancario ya esta conciliado en este run.",
                )
            if match.ledger_transaction_id == ledger_transaction_id:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="El movimiento contable ya esta conciliado en este run.",
                )

    def _recalculate_run_counts(self, run: ReconciliationRun) -> ReconciliationRun:
        matches = list(self.match_repository.list_by_run_id(run.id))
        run.matched_count = sum(1 for match in matches if match.status == MATCHED)
        run.unmatched_bank_count = sum(1 for match in matches if match.status == UNMATCHED_BANK)
        run.unmatched_ledger_count = sum(1 for match in matches if match.status == UNMATCHED_LEDGER)
        run.possible_match_count = sum(1 for match in matches if match.status == POSSIBLE_MATCH)
        run.total_bank = run.matched_count + run.unmatched_bank_count + run.possible_match_count
        run.total_ledger = run.matched_count + run.unmatched_ledger_count + run.possible_match_count
        return self.run_repository.save(run)

    def _populate_summary_sheet(self, sheet, run: ReconciliationRun) -> None:
        rows = [
            ("Run ID", str(run.id)),
            ("Estado", run.status),
            ("Fecha creación", run.created_at.isoformat()),
            ("Total Banco", run.total_bank),
            ("Total Contabilidad", run.total_ledger),
            ("Conciliados", run.matched_count),
            ("Posibles", run.possible_match_count),
            ("No conciliados banco", run.unmatched_bank_count),
            ("No conciliados contabilidad", run.unmatched_ledger_count),
        ]
        for row in rows:
            sheet.append(row)
        self._autosize_columns(sheet)

    def _populate_matched_sheet(
        self,
        sheet,
        matches: list[ReconciliationMatch],
        bank_transactions: dict[uuid.UUID, BankTransaction],
        ledger_transactions: dict[uuid.UUID, LedgerTransaction],
    ) -> None:
        sheet.append(
            [
                "Banco ID",
                "Contabilidad ID",
                "Monto Banco",
                "Monto Contabilidad",
                "Diferencia",
                "Días diferencia",
                "Similitud descripción",
                "Estado",
                "Notas",
            ]
        )
        for match in matches:
            if match.status not in {MATCHED, POSSIBLE_MATCH}:
                continue
            bank_transaction = (
                bank_transactions.get(match.bank_transaction_id) if match.bank_transaction_id is not None else None
            )
            ledger_transaction = (
                ledger_transactions.get(match.ledger_transaction_id) if match.ledger_transaction_id is not None else None
            )
            sheet.append(
                [
                    str(match.bank_transaction_id) if match.bank_transaction_id else "",
                    str(match.ledger_transaction_id) if match.ledger_transaction_id else "",
                    bank_transaction.amount if bank_transaction is not None else "",
                    self._ledger_amount(ledger_transaction) if ledger_transaction is not None else "",
                    match.amount_difference or "",
                    match.days_difference if match.days_difference is not None else "",
                    match.description_similarity if match.description_similarity is not None else "",
                    match.status,
                    match.notes or "",
                ]
            )
        self._autosize_columns(sheet)

    def _populate_unmatched_bank_sheet(
        self,
        sheet,
        matches: list[ReconciliationMatch],
        bank_transactions: dict[uuid.UUID, BankTransaction],
    ) -> None:
        sheet.append(["Banco ID", "Fecha", "Descripción", "Monto", "Notas"])
        for match in matches:
            if match.status != UNMATCHED_BANK or match.bank_transaction_id is None:
                continue
            bank_transaction = bank_transactions.get(match.bank_transaction_id)
            if bank_transaction is None:
                continue
            sheet.append(
                [
                    str(bank_transaction.id),
                    bank_transaction.transaction_date.isoformat(),
                    bank_transaction.description,
                    bank_transaction.amount,
                    match.notes or "",
                ]
            )
        self._autosize_columns(sheet)

    def _populate_unmatched_ledger_sheet(
        self,
        sheet,
        matches: list[ReconciliationMatch],
        ledger_transactions: dict[uuid.UUID, LedgerTransaction],
    ) -> None:
        sheet.append(["Contabilidad ID", "Fecha", "Descripción", "Débito", "Crédito", "Notas"])
        for match in matches:
            if match.status != UNMATCHED_LEDGER or match.ledger_transaction_id is None:
                continue
            ledger_transaction = ledger_transactions.get(match.ledger_transaction_id)
            if ledger_transaction is None:
                continue
            sheet.append(
                [
                    str(ledger_transaction.id),
                    ledger_transaction.transaction_date.isoformat(),
                    ledger_transaction.account_name,
                    ledger_transaction.debit,
                    ledger_transaction.credit,
                    match.notes or "",
                ]
            )
        self._autosize_columns(sheet)

    @staticmethod
    def _autosize_columns(sheet) -> None:
        for column_cells in sheet.columns:
            values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
            width = max(values, default=0) + 2
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = min(width, 60)

    def _build_matches(
        self,
        *,
        bank_transactions: list[BankTransaction],
        ledger_transactions: list[LedgerTransaction],
        tolerance_amount: Decimal,
        tolerance_days: int,
    ) -> tuple[list[MatchDraft], dict[str, int]]:
        candidates = self._collect_candidates(
            bank_transactions=bank_transactions,
            ledger_transactions=ledger_transactions,
            tolerance_amount=tolerance_amount,
            tolerance_days=tolerance_days,
        )
        candidates.sort(
            key=lambda candidate: (
                candidate.priority,
                candidate.amount_difference,
                candidate.days_difference,
                -candidate.description_similarity,
                str(candidate.bank_transaction.id),
                str(candidate.ledger_transaction.id),
            )
        )

        used_bank_ids: set[uuid.UUID] = set()
        used_ledger_ids: set[uuid.UUID] = set()
        results: list[MatchDraft] = []
        matched_count = 0
        unmatched_bank_count = 0
        possible_match_count = 0

        for candidate in candidates:
            if candidate.bank_transaction.id in used_bank_ids:
                continue
            if candidate.ledger_transaction.id in used_ledger_ids:
                continue

            results.append(
                MatchDraft(
                    bank_transaction_id=candidate.bank_transaction.id,
                    ledger_transaction_id=candidate.ledger_transaction.id,
                    status=candidate.status,
                    match_score=candidate.match_score,
                    amount_difference=candidate.amount_difference,
                    days_difference=candidate.days_difference,
                    description_similarity=candidate.description_similarity,
                    notes=candidate.notes,
                )
            )
            used_bank_ids.add(candidate.bank_transaction.id)
            used_ledger_ids.add(candidate.ledger_transaction.id)
            if candidate.status == MATCHED:
                matched_count += 1
            elif candidate.status == POSSIBLE_MATCH:
                possible_match_count += 1

        unmatched_ledger_count = 0
        for bank_transaction in bank_transactions:
            if bank_transaction.id in used_bank_ids:
                continue
            results.append(
                MatchDraft(
                    bank_transaction_id=bank_transaction.id,
                    ledger_transaction_id=None,
                    status=UNMATCHED_BANK,
                    match_score=0,
                    amount_difference=None,
                    days_difference=None,
                    description_similarity=None,
                    notes="Sin candidato contable dentro de la tolerancia.",
                )
            )
            unmatched_bank_count += 1

        for ledger_transaction in ledger_transactions:
            if ledger_transaction.id in used_ledger_ids:
                continue
            results.append(
                MatchDraft(
                    bank_transaction_id=None,
                    ledger_transaction_id=ledger_transaction.id,
                    status=UNMATCHED_LEDGER,
                    match_score=0,
                    amount_difference=None,
                    days_difference=None,
                    description_similarity=None,
                    notes="Sin movimiento bancario equivalente dentro de la tolerancia.",
                )
            )
            unmatched_ledger_count += 1

        summary = {
            "total_bank": len(bank_transactions),
            "total_ledger": len(ledger_transactions),
            "matched_count": matched_count,
            "unmatched_bank_count": unmatched_bank_count,
            "unmatched_ledger_count": unmatched_ledger_count,
            "possible_match_count": possible_match_count,
        }
        return results, summary

    def _collect_candidates(
        self,
        *,
        bank_transactions: list[BankTransaction],
        ledger_transactions: list[LedgerTransaction],
        tolerance_amount: Decimal,
        tolerance_days: int,
    ) -> list[CandidateMatch]:
        candidates: list[CandidateMatch] = []
        for bank_transaction in bank_transactions:
            for ledger_transaction in ledger_transactions:
                candidate = self._build_candidate(
                    bank_transaction=bank_transaction,
                    ledger_transaction=ledger_transaction,
                    tolerance_amount=tolerance_amount,
                    tolerance_days=tolerance_days,
                )
                if candidate is not None:
                    candidates.append(candidate)
        return candidates

    def _build_candidate(
        self,
        *,
        bank_transaction: BankTransaction,
        ledger_transaction: LedgerTransaction,
        tolerance_amount: Decimal,
        tolerance_days: int,
    ) -> CandidateMatch | None:
        ledger_amount = self._ledger_amount(ledger_transaction)
        amount_difference = abs(bank_transaction.amount - ledger_amount)
        days_difference = abs((bank_transaction.transaction_date - ledger_transaction.transaction_date).days)
        similarity = self._description_similarity(bank_transaction, ledger_transaction)

        amount_exact = amount_difference == Decimal("0.00")
        date_exact = days_difference == 0
        amount_within_tolerance = amount_difference <= tolerance_amount
        date_within_tolerance = days_difference <= tolerance_days
        partial_similarity = Decimal("0.0000") < similarity < DESCRIPTION_MATCH_THRESHOLD
        possible_amount_limit = max(tolerance_amount, POSSIBLE_MATCH_MAX_AMOUNT_DIFFERENCE)
        possible_days_limit = max(tolerance_days, POSSIBLE_MATCH_MAX_DAYS_DIFFERENCE)

        if amount_exact and date_exact:
            return CandidateMatch(
                bank_transaction=bank_transaction,
                ledger_transaction=ledger_transaction,
                status=MATCHED,
                amount_difference=amount_difference,
                days_difference=days_difference,
                description_similarity=similarity,
                match_score=self._build_match_score(
                    amount_difference=amount_difference,
                    days_difference=days_difference,
                    similarity=similarity,
                ),
                priority=0,
                notes=self._build_matched_note(
                    amount_exact=amount_exact,
                    description_similarity=similarity,
                ),
            )

        if amount_exact and date_within_tolerance:
            return CandidateMatch(
                bank_transaction=bank_transaction,
                ledger_transaction=ledger_transaction,
                status=MATCHED,
                amount_difference=amount_difference,
                days_difference=days_difference,
                description_similarity=similarity,
                match_score=self._build_match_score(
                    amount_difference=amount_difference,
                    days_difference=days_difference,
                    similarity=similarity,
                ),
                priority=1,
                notes=self._build_matched_note(
                    amount_exact=amount_exact,
                    description_similarity=similarity,
                ),
            )

        if amount_within_tolerance and date_within_tolerance:
            return CandidateMatch(
                bank_transaction=bank_transaction,
                ledger_transaction=ledger_transaction,
                status=MATCHED,
                amount_difference=amount_difference,
                days_difference=days_difference,
                description_similarity=similarity,
                match_score=self._build_match_score(
                    amount_difference=amount_difference,
                    days_difference=days_difference,
                    similarity=similarity,
                ),
                priority=2,
                notes="Conciliado por tolerancia de valor y fecha.",
            )

        if (
            amount_difference > Decimal("0.00")
            and amount_difference <= possible_amount_limit
            and days_difference > 0
            and days_difference <= possible_days_limit
            and partial_similarity
        ):
            return CandidateMatch(
                bank_transaction=bank_transaction,
                ledger_transaction=ledger_transaction,
                status=POSSIBLE_MATCH,
                amount_difference=amount_difference,
                days_difference=days_difference,
                description_similarity=similarity,
                match_score=self._build_match_score(
                    amount_difference=amount_difference,
                    days_difference=days_difference,
                    similarity=similarity,
                ),
                priority=3,
                notes="Se encontró una coincidencia cercana por valor, fecha y similitud parcial.",
            )

        return None

    @staticmethod
    def _ledger_amount(transaction: LedgerTransaction) -> Decimal:
        return (transaction.debit - transaction.credit).quantize(Decimal("0.01"))

    def _description_similarity(
        self,
        bank_transaction: BankTransaction,
        ledger_transaction: LedgerTransaction,
    ) -> Decimal:
        bank_text = self._normalize_for_similarity(
            bank_transaction.description,
            bank_transaction.reference,
            bank_transaction.document_number,
        )
        ledger_text = self._normalize_for_similarity(
            ledger_transaction.account_name,
            ledger_transaction.third_party,
            ledger_transaction.document_number,
        )
        if not bank_text or not ledger_text:
            return Decimal("0.0000")
        ratio = SequenceMatcher(None, bank_text, ledger_text).ratio()
        return Decimal(f"{ratio:.4f}")

    @staticmethod
    def _normalize_for_similarity(*parts: str | None) -> str:
        tokens: list[str] = []
        for part in parts:
            if not part:
                continue
            normalized = re.sub(r"[^a-z0-9\s]", " ", part.lower())
            normalized = re.sub(r"\s+", " ", normalized).strip()
            if normalized:
                tokens.append(normalized)
        return " ".join(tokens)

    @staticmethod
    def _build_match_score(
        *,
        amount_difference: Decimal,
        days_difference: int,
        similarity: Decimal,
    ) -> int:
        score = 100
        score -= int(amount_difference * 10)
        score -= days_difference * 5
        score += int(similarity * 10)
        return max(score, 0)

    @staticmethod
    def _build_matched_note(
        *,
        amount_exact: bool,
        description_similarity: Decimal,
    ) -> str:
        if amount_exact and description_similarity < DESCRIPTION_MATCH_THRESHOLD:
            return "Conciliado por valor y fecha; descripción diferente."
        return "Coincidencia automática por valor y fecha."
